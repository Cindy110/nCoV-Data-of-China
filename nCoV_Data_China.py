'''
Created on 2020-02-18
author: Cindy110
Function:   Create a epidemic report of China which includes the number of "Confirmed", "suspect", "Dead", etc.
            Create a traffic information form of infected people
'''

# For Crawler
import json
import requests
from bs4 import BeautifulSoup as BS
import re
from datetime import datetime
# For save data
import pandas as pd
import numpy as np
import openpyxl

# Data Source:https://news.qq.com/zt2020/page/feiyan.htm#/global
url_Report_Data = 'https://view.inews.qq.com/g2/getOnsInfo?name=disease_h5'
url_Traffic_Data = 'http://hhyfeed.sogoucdn.com/js/common/epidemic-search/main_2020020220.js'

def HTMLGet(url):
    header = {'User-Agent':
                  "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.102 Safari/537.36 Edge/18.18362"
              }
    try:
        r = requests.get(url, headers=header)
        r.encoding = r.apparent_encoding
        return r.text
    except:
        return "Error"

def DataGet(htmlr):
    # Data of epidemic surveillance
    area = []
    dead = []
    deadRate = []
    healRate = []
    confirm = []
    suspect = []
    heal = []
    # Data of transportation surveillance
    Date = []
    Ticket_number = []
    start_point = []
    end_point = []

    # json -> class dict
    DataReport_CN_temp = json.loads(htmlr)
    DataReport_CN_temp = DataReport_CN_temp["data"]
    # 正则表达式的 rule
    name_rule = r'"name":[^,]+'
    name_rule2 = r'[\u4E00-\u9FA5]+'    # Chainese
    data_rule = r'{[^{]+"showRate"[^}]+}'
    confirm_rule = r'"confirm":[^,]+'
    suspect_rule = r'"suspect":[^,]+'
    dead_rule = r'"dead":[^,]+'
    heal_rule = r'"heal":[^,]+'
    deadRate_rule = r'"deadRate":[^,]+'
    healRate_rule = r'"healRate":[^,]+'
    count_rule = r'\d+'
    float_rule = r'\d{1,3}\.\d{2}'

    Name_list = re.findall(name_rule,DataReport_CN_temp)
    Data_list = re.findall(data_rule,DataReport_CN_temp)
    Data_list = "".join(Data_list)
    # obtain all data list
    Confirm_list = re.findall(confirm_rule,Data_list)
    Suspect_list = re.findall(suspect_rule, Data_list)
    Dead_list = re.findall(dead_rule, Data_list)
    Heal_list = re.findall(heal_rule, Data_list)
    DeadRate_list = re.findall(deadRate_rule, Data_list)
    HealRate_list = re.findall(healRate_rule, Data_list)

    for i in range(0,len(Name_list)):
        # 城市名称
        Name_temp = re.findall(name_rule2,Name_list[i])
        Name_temp = "".join(Name_temp)  # list changes to str
        area.append(Name_temp)
        # 确诊人数
        Confirm_temp = re.findall(count_rule, Confirm_list[i])
        Confirm_temp = "".join(Confirm_temp)
        confirm.append(Confirm_temp)
        # 疑似病例
        Suspect_temp = re.findall(count_rule, Suspect_list[i])
        Suspect_temp = "".join(Suspect_temp)
        suspect.append(Suspect_temp)
        # 死亡人数
        Dead_temp = re.findall(count_rule, Dead_list[i])
        Dead_temp = "".join(Dead_temp)
        dead.append(Dead_temp)
        # 痊愈人数
        Heal_temp = re.findall(count_rule, Heal_list[i])
        Heal_temp = "".join(Heal_temp)
        heal.append(Heal_temp)
        # 死亡率
        DeadRate_temp = re.findall(float_rule, DeadRate_list[i])
        DeadRate_temp = "".join(DeadRate_temp)
        deadRate.append(DeadRate_temp)
        # 痊愈率
        HealRate_temp = re.findall(float_rule, HealRate_list[i])
        HealRate_temp = "".join(HealRate_temp)
        healRate.append(HealRate_temp)

        # Save data
        # Create workbook for report data
        wb = openpyxl.Workbook()
        ws = wb.active
        # write into cell
        ws.cell(1, 1, value='Area')
        ws.cell(1, 2, value='Confirm')
        ws.cell(1, 3, value='Suspect')
        ws.cell(1, 4, value='Heal')
        ws.cell(1, 5, value='Dead')
        ws.cell(1, 6, value='Heal Rate')
        ws.cell(1, 7, value='Dead Rate')

        for j in range(0,len(area)):
            ws.cell(j + 2, 1, value=area[j])
            ws.cell(j + 2, 2, value=confirm[j])
            ws.cell(j + 2, 3, value=suspect[j])
            ws.cell(j + 2, 4, value=heal[j])
            ws.cell(j + 2, 5, value=dead[j])
            ws.cell(j + 2, 6, value=healRate[j])
            ws.cell(j + 2, 7, value=deadRate[j])
        wb.save('Epidemic Report.xlsx')   # Epidemic Report

    return "Epidemic Report is Created"

def traffic_info(html):
    traffic_type = []
    traffic_num = []
    traffic_path = []
    traffic_time = []
    update_time = []
    traffic_info_temp = json.loads(html)    # class 'list'
    # obtain traffic info
    for dict_element in traffic_info_temp:
        traffic_type.append(dict_element['trafficType'])
        traffic_num.append(dict_element['trafficNum'])
        traffic_path.append(dict_element['trainPath'])
        traffic_time.append(dict_element['trafficTime'])
        update_time.append(dict_element['updateTime'])
    # obtain update time
    for updateTime_i in update_time:
        if updateTime_i:
            updateTime_temp = updateTime_i
            break

    wb = openpyxl.Workbook()
    ws = wb.active
    # Update Time
    ws.cell(1, 1, value='Update Time')
    ws.cell(1, 2, value=updateTime_temp)
    # write into cell
    for i in range(0,len(traffic_type)):
        ws.cell(i + 2, 1, value=traffic_type[i])
        ws.cell(i + 2, 2, value=traffic_num[i])
        ws.cell(i + 2, 3, value=traffic_time[i])
        ws.cell(i + 2, 4, value=traffic_path[i])
    wb.save('Traffic Information Form.xlsx')

    return "Traffic Information Form is Created"

def main():

    Traffic_info = HTMLGet(url_Traffic_Data)
    Report_info = HTMLGet(url_Report_Data)
    # Creating forms
    print(traffic_info(Traffic_info))
    print(DataGet(Report_info))

if __name__ == '__main__':
    main()